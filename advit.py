from netmiko import ConnectHandler
from netmiko.ssh_exception import AuthenticationException
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors , PatternFill
import re
from datetime import datetime

workbook = Workbook()
def Ping(sheet,sheet_title,ping_ip,num,file_name):

    #workbook = Workbook()
    #sheet = workbook.active

    #workbook = Workbook()
    #sheet = workbook.worksheets[int(num)]
    #sheet = workbook.active 
    sheet = workbook.create_sheet(sheet_title,int(num))
    
    


 
    #wb = Workbook()

    #sheet = workbook.create_sheet("Sheet_A",int(num))
    #sheet.title = sheet_title

    #sheet = workbook.create_sheet("Sheet_B", 1)
    #sheet.title = sheet_title

    #wb.save(filename = 'sample_book.xlsx')

    sheet["A1"] = "Management IP"
    sheet["B1"] = "Hostname"
    sheet["C1"] = "GBN ID"
    sheet["D1"] = "Model"
    sheet["E1"] = "Ping Status"
    sheet["F1"] = "Success Rate"
    sheet["G1"] = "LAN Subnet"
    sheet["H1"] = "LAN Status"
    sheet["I1"] = "Date/Time"



    username="gbn_n"
    password="Admin@123"

    def Date_time(start_row,start_column):
                
        dt_object = datetime.now()

        sheet.cell(row=start_row, column=start_column).value = dt_object

        return

    def Background_color(row,color,id1,id2):
        fill = PatternFill(start_color=color, fill_type = "solid")


        for cell in sheet[row][id1:id2]:
            cell.fill = fill


        
        return 

    Background_color(1,"009999FF",0,9)

    with open(file_name) as f:
        endpointIPs = f.readlines()

    start_row = 2
    start_column = 1

    for endpoint in endpointIPs:
        host = endpoint.strip()
        print(host)
        print ('Connecting to device: ' + host)
        ios_device = {
            'device_type': 'cisco_ios',
            'ip': host,
            'username': username,
            'password': password
        }

        sheet.cell(row=start_row, column=start_column).value = host
        start_column += 1
        print(start_column)
        print(start_row)

        try:

                    
            remote_conn = ConnectHandler(**ios_device)

            #sheet.cell(row=start_row, column=start_column).value = host
            #start_column += 1
            #print(start_column)
            #print(start_row)

            GBN_ID= remote_conn.send_command("dis current-configuration | i sysname ")
            print (GBN_ID)

            hostname = GBN_ID.split(" ")[2]
            print(hostname)

            sheet.cell(row=start_row, column=start_column).value = hostname
            start_column += 1
            
            print(hostname[0:8])
            sheet.cell(row=start_row, column=start_column).value = hostname[0:8]
            start_column += 1

            sheet.cell(row=start_row, column=start_column).value = hostname[21:25]
            start_column += 1
            
            output= remote_conn.send_command("dis ip int brief ")
            print (output)
        

            
            match = re.compile(r'(10)\.(\d|\d\d|\d\d\d)\.(\d|\d\d|\d\d\d)\.(\d|\d\d|\d\d\d)')
            match1 = re.compile(r'(10)\.(\d|\d\d|\d\d\d)\.(\d|\d\d|\d\d\d)\.(\d|\d\d|\d\d\d)/(\d\d)\s*(up|down|\*down)')
            
            
            matches = match.finditer(output)
            matches1 = match1.finditer(output)
            

            
            #print(matches1)

            # matches = pattern.finditer(urls)
            List=[]
            for match in matches:
               #print(match)
               List.append(match[0])
            print(List[0])

            List1=[]
            for match1 in matches1:
                #print(match2)
                List1.append(match1[0])
            print(List1)

            #if List1.__contains__("up"):
                

            
            status=[]
            for i in List1:
                status.append(i.split(' '))
            #print(status)


     

            
            #print("Interface is Up")
            for item in List:
                print(item)
                output1= remote_conn.send_command("ping -a %s %s" % (item,ping_ip) )
                print("ping -a %s %s" % (item," 8.8.8.8"))
                print (output1)

                

                match2 = re.compile(r'(0|25|50|75|100)\.(00\%)')

                matches2 = match2.finditer(output1)

                List2 = []
                for match2 in matches2:
                    # print(match)
                    List2.append(match2[0])
                    #print(match2[0])
                    print(List2)
                    print("advit")

                
               
                for id in List2:

                    
                    if id == "0.00%":
                        sheet.cell(row=start_row, column=start_column).value = "Pingable"

                        Background_color(start_row,"00008000",4,5)
                        
                        start_column += 1

                        id1 = 100
                                                
                        #sheet.cell(row=start_row, column=start_column).value = "=MIN(100-20)"
                        sheet.cell(row=start_row, column=start_column).value = id1
                        #print(start_column)

                        #Background_color(start_row,"00008000",7,8)

                        start_column += 1
                        print(start_column)
                        print(start_row)
                                            


                    elif id == "25.00%":
                        sheet.cell(row=start_row, column=start_column).value = "Pingable"

                        Background_color(start_row,"00008000",4,5)
                        
                        start_column += 1

                        id1 = 75
                                                
                        #sheet.cell(row=start_row, column=start_column).value = "=MIN(100-20)"
                        sheet.cell(row=start_row, column=start_column).value = id1
                        #print(start_column)
                        
                        #start_row += 1
                        start_column += 1
                        print(start_column)
                        print(start_row)

                    elif id == "50.00%":
                        sheet.cell(row=start_row, column=start_column).value = "Pingable"

                        Background_color(start_row,"00008000",4,5)
                        
                        start_column += 1

                        id1 = 50
                                                
                        #sheet.cell(row=start_row, column=start_column).value = "=MIN(100-20)"
                        sheet.cell(row=start_row, column=start_column).value = id1
                        #print(start_column)
                        
                        #start_row += 1
                        start_column += 1
                        print(start_column)
                        print(start_row)


                    elif id == "75.00%":
                        sheet.cell(row=start_row, column=start_column).value = "Pingable"

                        Background_color(start_row,"00008000",4,5)
                        
                        start_column += 1

                        id1 = 25
                                                
                        #sheet.cell(row=start_row, column=start_column).value = "=MIN(100-20)"
                        sheet.cell(row=start_row, column=start_column).value = id1
                        #print(start_column)
                        
                        #start_row += 1
                        start_column += 1
                        print(start_column)
                        print(start_row)
                        

                    else:
                        sheet.cell(row=start_row, column=start_column).value = "Not Pingable"

                        Background_color(start_row,"00FF0000",4,5)
                        
                        start_column += 1
                        sheet.cell(row=start_row, column=start_column).value = "Unsuccessfull"

                        #print(start_column)
                        
                        #start_row += 1
                        start_column += 1
                        print(start_column)
                        print(start_row)
                            

                                          
            

           

            print("============================================================\n\n")

            
            

            for ip in List:
                Lan = i.split(" ")
                print(Lan)

                sheet.cell(row=start_row, column=start_column).value = Lan[0]
            
                         
                #start_row += 1
                start_column += 1
                print(start_column)
                print(start_row)

                if Lan.__contains__("up"):
                    sheet.cell(row=start_row, column=start_column).value = "UP"

                    Background_color(start_row,"00008000",7,8)

                    start_column += 1
                    
                    #start_row += 1
                    #start_column -= 7
                    

                else:
                    sheet.cell(row=start_row, column=start_column).value = "Down"

                    Background_color(start_row,"00FF0000",7,8)

                    start_column += 1
                    #start_row += 1
                    #start_column -= 7

        

            

            Date_time(start_row,start_column)
            
            start_row += 1
            start_column -= 8

                    

        except (AuthenticationException):
            print ('Authentication failure: ' + host)
            sheet.merge_cells(start_row=start_row,start_column=start_column,end_row=start_row,end_column=8)
            sheet.cell(row=start_row, column=start_column).value = "Authentication failure"
            

            Background_color(start_row,"00800000",1,2)
            
            start_column += 1     

            
            Date_time(start_row,9)

           
            start_row += 1
            start_column -= 2
            print(start_column)
            print(start_row)

            print("============================================================\n\n")
                   
            continue
       
        except (EOFError):
            print ('End of file while attempting device ' + host)
            sheet.merge_cells(start_row=start_row,start_column=start_column,end_row=start_row,end_column=8)
            sheet.cell(row=start_row, column=start_column).value = "End of file while attempting device"
            
            Background_color(start_row,"00800000",1,2)
            
            start_column += 1     

            
            Date_time(start_row,9)

           
            start_row += 1
            start_column -= 2
            
            print(start_column)
            print(start_row)
            
            print("============================================================\n\n")
            continue
       
        except:
            print ('Device is not reachable')
            sheet.merge_cells(start_row=start_row,start_column=start_column,end_row=start_row,end_column=8)
            
            sheet.cell(row=start_row, column=start_column).value = "Device is not reachable"

            Background_color(start_row,"00FFFF00",1,2)

            start_column += 1     

            
            Date_time(start_row,9)

           
            start_row += 1
            start_column -= 2
            
            print(start_column)
            print(start_row)
            print("============================================================\n\n")
            continue


    # Create a few styles
    font = Font(bold=True,
                    size=11,
                    color="00000000")

    alignment=Alignment(horizontal='center',
                              vertical='center',
                              text_rotation=0,
                              wrap_text=False,
                              shrink_to_fit=False,
                              indent=0)

    border = Border(left=Side(border_style="thin",
                                     color='FF000000'),
                            right=Side(border_style="thin",
                                      color='FF000000'),
                            top=Side(border_style="thin",
                                     color='FF000000'),
                            bottom=Side(border_style="thin",
                                        color='FF000000'),
                            diagonal=Side(border_style="thin",
                                          color='FF000000'),
                            diagonal_direction=0,
                            outline=Side(border_style="thin",
                                         color='FF000000'),
                            vertical=Side(border_style="thin",
                                          color='FF000000'),
                            horizontal=Side(border_style="thin",
                                           color='FF000000')
                           )

    #fill = PatternFill(start_color="00008080", end_color="00008080", fill_type = "solid")

    for cell in sheet[1:1]:
        cell.font = font

            
    print(start_row)
    for i in range(1,start_row):
        for cell in sheet[i]:
            cell.alignment = alignment
            cell.border = border
            #cell.fill = fill

        

    #sheet.title = sheet_title
    print(workbook.sheetnames)
    workbook.save(filename="GBN_Ping.xlsx")

    return

Ping("wb1","GBN_Ping_Data","8.8.8.8","0",'endpointIPs_1.txt')
Ping("wb2","GPG_Ping_Data","10.142.200.1","1",'endpointIPs_2.txt')
